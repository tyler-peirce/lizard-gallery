#!/usr/bin/env python3
"""
sync_from_spreadsheet.py
========================
Rebuild animals.json from the catalogue spreadsheet so the website matches
the spreadsheet. Run this after editing the spreadsheet.

WHAT IT DOES
  - Reads the catalogue .xlsx (Sheet1).
  - For every animal that is FOR SALE (numeric "Rough_price", not "Sold"):
      * price          <- Rough_price
      * sex            <- Sex            (M / F / blank)
      * dob            <- Birth_date     (YYYY-MM-DD)
      * breeding_group <- "Breeding group" column (e.g. Q1); cleared if sold
      * morphs         <- every morph column whose value is "Visual"
      * hets           <- every morph column whose value is "100% het"
      * notes          <- visual morphs, then the remaining het percentages
                          (e.g. "T+, 50% het Anery, pos het Hyper")
  - Removes animals marked "Sold" in the spreadsheet.
  - Adds NEW for-sale animals that aren't in animals.json yet, but ONLY when a
    photos/<id>/ folder already exists (so we never create a broken card).
  - Rebuilds the top-level "breeding_groups" block (price + description),
    excluding any sold members.

WHAT IT PRESERVES (never derivable from columns)
  - Hand-typed free-text notes when the columns produce no note
    (e.g. "Wild (Prince Regent Local)").
  - The "config" block and any breeding-group description the spreadsheet
    leaves blank (falls back to the existing description).

USAGE
  python3 sync_from_spreadsheet.py            # apply changes, print a report
  python3 sync_from_spreadsheet.py --dry-run  # preview only, write nothing
  python3 sync_from_spreadsheet.py --sync-photos   # also run ./sync-photos.sh

Requires openpyxl:  pip install openpyxl
"""
import argparse
import datetime as dt
import json
import os
import subprocess
import sys

# ── Column -> website morph label ───────────────────────────────────────────
# Order matters: it controls the order traits appear in the "notes" string.
COLMAP = [
    ("Albino", "Albino"),
    ("Anery", "Anery"),
    ("T+_albino", "T+"),
    ("White_norhtern", "White Northern"),   # header is misspelled in the sheet
    ("Hyper", "Hyper"),
    ("Piedsided", "Piedsided"),
    ("Patchwork", "Patchwork"),
    ("Whitesided", "Whitesided"),
]

# Canonical key order for each animal object in animals.json.
KEY_ORDER = ["id", "breeding_group", "morphs", "hets", "price", "dob", "notes", "sex"]

HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_XLSX = os.path.join(HERE, "Catalogue-of-Lizards_Tyler-Peirce.xlsx")
ANIMALS_JSON = os.path.join(HERE, "animals.json")
PHOTOS_DIR = os.path.join(HERE, "photos")


def cell(value):
    """Normalise a cell value: strip strings, leave everything else as-is."""
    return value.strip() if isinstance(value, str) else value


def is_sold(row):
    rough = str(cell(row.get("Rough_price"))).lower()
    whole = str(cell(row.get("Wholesale_price_offer"))).lower()
    return rough == "sold" or whole == "sold"


def fmt_dob(value):
    """Birth_date -> 'YYYY-MM-DD'. Falls back to a plain year string."""
    if value is None:
        return None
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def derive_genetics(row):
    """Return (morphs, hets, notes) derived purely from the morph columns."""
    vals = {col: cell(row.get(col)) for col, _ in COLMAP}

    morphs = [label for col, label in COLMAP if vals[col] == "Visual"]
    hets = [label for col, label in COLMAP if vals[col] == "100% het"]

    # notes: visual morphs first (bare), then every other het percentage.
    # "100% het" traits live in `hets` (rendered as het badges), so they are
    # intentionally excluded from the notes string.
    parts = [label for col, label in COLMAP if vals[col] == "Visual"]
    for col, label in COLMAP:
        v = vals[col]
        if v in (None, "", "Visual", "100% het"):
            continue
        # Cell already reads like "50% het" / "66% het"; "pos" needs "het" added.
        parts.append(f"pos het {label}" if v == "pos" else f"{v} {label}")

    notes = ", ".join(parts) if parts else None
    return (morphs or None), (hets or None), notes


def ordered_animal(a):
    """Return a new dict with canonical key order, dropping empty optionals."""
    out = {}
    for k in KEY_ORDER:
        v = a.get(k)
        if v in (None, [], ""):
            continue
        out[k] = v
    # keep any unexpected custom keys (e.g. litter) at the end, untouched
    for k, v in a.items():
        if k not in out and v not in (None, [], ""):
            out[k] = v
    return out


def load_sheet(xlsx_path):
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl is required:  pip install openpyxl")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Sheet1"]
    headers = [c.value for c in ws[1]]
    rows = {}
    for raw in ws.iter_rows(min_row=2, values_only=True):
        record = {h: raw[i] for i, h in enumerate(headers) if h}
        code = cell(record.get("Animal_code"))
        if code:
            rows[str(code)] = record
    return rows


def main():
    ap = argparse.ArgumentParser(description="Sync animals.json from the catalogue spreadsheet.")
    ap.add_argument("xlsx", nargs="?", default=DEFAULT_XLSX, help="Path to the catalogue .xlsx")
    ap.add_argument("--dry-run", action="store_true", help="Preview changes without writing")
    ap.add_argument("--sync-photos", action="store_true", help="Run ./sync-photos.sh afterwards")
    args = ap.parse_args()

    sheet = load_sheet(args.xlsx)
    data = json.load(open(ANIMALS_JSON))
    existing = {a["id"]: a for a in data["animals"]}
    existing_order = [a["id"] for a in data["animals"]]

    report = {"removed": [], "added": [], "skipped_no_photo": [],
              "price": [], "sex": [], "bg": [], "notes": [], "poa": [],
              "morphs": [], "hets": [], "dob": [], "groups": []}

    def build(aid, prev):
        """Build a fully-synced animal dict from sheet row `aid`, using `prev`
        (existing json entry, possibly {}) to preserve free-text notes."""
        row = sheet[aid]
        morphs, hets, derived_notes = derive_genetics(row)

        rough = cell(row.get("Rough_price"))
        price = int(rough) if isinstance(rough, (int, float)) else None
        if price is None:
            report["poa"].append(aid)

        sex = cell(row.get("Sex")) or None
        dob = fmt_dob(cell(row.get("Birth_date")))
        bg = (cell(row.get("Breeding group")) or None) if not is_sold(row) else None

        # Preserve a hand-typed note when the columns yield nothing derivable.
        notes = derived_notes
        if notes is None and prev.get("notes"):
            notes = prev["notes"]

        # Change tracking (vs previous json) for the report
        if prev:
            if prev.get("price") != price:
                report["price"].append((aid, prev.get("price"), price))
            if (prev.get("sex") or None) != sex:
                report["sex"].append((aid, prev.get("sex"), sex))
            if (prev.get("breeding_group") or None) != bg:
                report["bg"].append((aid, prev.get("breeding_group"), bg))
            if (prev.get("notes") or None) != (notes or None):
                report["notes"].append((aid, prev.get("notes"), notes))
            if (prev.get("morphs") or None) != morphs:
                report["morphs"].append((aid, prev.get("morphs"), morphs))
            if (prev.get("hets") or None) != hets:
                report["hets"].append((aid, prev.get("hets"), hets))
            if (prev.get("dob") or None) != (dob or None):
                report["dob"].append((aid, prev.get("dob"), dob))

        return ordered_animal({
            "id": aid, "breeding_group": bg, "morphs": morphs, "hets": hets,
            "price": price, "dob": dob, "notes": notes, "sex": sex,
        })

    new_animals = []

    # 1) Walk existing animals in their current order: keep, update, or drop.
    for aid in existing_order:
        row = sheet.get(aid)
        if row is None:
            # Not in the spreadsheet at all -> keep as-is (manually maintained).
            new_animals.append(ordered_animal(existing[aid]))
            continue
        if is_sold(row):
            report["removed"].append(aid)
            continue
        new_animals.append(build(aid, existing[aid]))

    # 2) Add NEW for-sale animals from the sheet (numeric price, has a photo folder).
    for aid, row in sheet.items():
        if aid in existing or is_sold(row):
            continue
        rough = cell(row.get("Rough_price"))
        if not isinstance(rough, (int, float)):
            continue  # blank / non-numeric and not sold -> not a listing
        if os.path.isdir(os.path.join(PHOTOS_DIR, aid)):
            new_animals.append(build(aid, {}))
            report["added"].append(aid)
        else:
            report["skipped_no_photo"].append(aid)

    # 3) Rebuild the breeding_groups block from the sheet (exclude sold members).
    groups = {}        # id -> {price, members[]}
    for aid, row in sheet.items():
        bg = cell(row.get("Breeding group"))
        if bg and not is_sold(row):
            grp = groups.setdefault(bg, {"price": None, "members": []})
            grp["members"].append(aid)
            gp = cell(row.get("Breeding group price"))
            if isinstance(gp, (int, float)):
                grp["price"] = int(gp)

    old_groups = data.get("breeding_groups", {})
    new_groups = {}
    for gid in sorted(groups):
        # description: first non-empty member cell, else keep the old one.
        desc = None
        for m in groups[gid]["members"]:
            d = cell(sheet[m].get("Breeding group description"))
            if d:
                desc = d
                break
        if not desc:
            desc = (old_groups.get(gid) or {}).get("description", "")
        new_groups[gid] = {"price": groups[gid]["price"], "description": desc}

    # Track breeding-group changes (added / removed / price / description).
    for gid in sorted(set(old_groups) | set(new_groups)):
        old, new = old_groups.get(gid), new_groups.get(gid)
        if old is None:
            report["groups"].append(f"{gid}: ADDED (${new['price']:,})")
        elif new is None:
            report["groups"].append(f"{gid}: REMOVED")
        else:
            if old.get("price") != new["price"]:
                report["groups"].append(f"{gid}: price {old.get('price')} -> {new['price']}")
            if (old.get("description") or "") != (new["description"] or ""):
                report["groups"].append(f"{gid}: description updated")

    # 4) Assemble output preserving top-level order: config, breeding_groups, animals.
    out = {}
    for k, v in data.items():
        if k in ("breeding_groups", "animals"):
            continue
        out[k] = v
    out["breeding_groups"] = new_groups
    out["animals"] = new_animals

    rendered = json.dumps(out, indent=2, ensure_ascii=False) + "\n"
    changed = rendered != open(ANIMALS_JSON).read()

    # ── Report ───────────────────────────────────────────────────────────────
    print(f"Spreadsheet: {os.path.basename(args.xlsx)}")
    print(f"Animals: {len(existing_order)} -> {len(new_animals)}")
    group_summary = ", ".join("{} (${:,})".format(g, new_groups[g]["price"]) for g in new_groups) or "none"
    print(f"Breeding groups: {group_summary}")

    def show(label, items, fmt):
        if items:
            print(f"\n{label} ({len(items)}):")
            for it in items:
                print("  " + fmt(it))

    show("Removed (sold)", report["removed"], lambda x: x)
    show("Added (new + photos found)", report["added"], lambda x: x)
    show("Skipped (for sale but NO photos/<id>/ folder)", report["skipped_no_photo"], lambda x: x)
    show("Price changes", report["price"], lambda x: f"{x[0]}: {x[1]} -> {x[2]}")
    show("Sex changes", report["sex"], lambda x: f"{x[0]}: {x[1]} -> {x[2]}")
    show("Animal group-membership changes", report["bg"], lambda x: f"{x[0]}: {x[1]} -> {x[2]}")
    show("Morph changes", report["morphs"], lambda x: f"{x[0]}: {x[1]} -> {x[2]}")
    show("Het changes", report["hets"], lambda x: f"{x[0]}: {x[1]} -> {x[2]}")
    show("DOB changes", report["dob"], lambda x: f"{x[0]}: {x[1]} -> {x[2]}")
    show("Note changes", report["notes"], lambda x: f"{x[0]}: {x[1]!r} -> {x[2]!r}")
    show("Breeding-group block changes", report["groups"], lambda x: x)
    show("Listed as POA (blank price)", report["poa"], lambda x: x)

    if not changed:
        print("\nanimals.json already matches the spreadsheet. No changes.")
        return

    if args.dry_run:
        print("\n[dry-run] animals.json WOULD change. Re-run without --dry-run to write.")
        return

    with open(ANIMALS_JSON, "w") as f:
        f.write(rendered)
    print("\nanimals.json updated.")

    if args.sync_photos:
        print("\nRunning ./sync-photos.sh ...")
        subprocess.run(["bash", os.path.join(HERE, "sync-photos.sh")], check=False)
    else:
        print("Tip: run ./sync-photos.sh to align git photo tracking, then commit & push.")


if __name__ == "__main__":
    main()
